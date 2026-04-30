"""
Genera reportes del último relevamiento en tres formatos:
  - Markdown (.md) – para leer rápido o pegar en WhatsApp/Slack
  - Excel    (.xlsx) – con dos hojas: Comparativa y Detalle
  - PDF      (.pdf)  – para imprimir o mandar adjunto

El reporte compara contra el relevamiento anterior y resalta variaciones >5%.
"""
import logging
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from statistics import mean

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
)

from normalizador import corte_pretty
from storage import obtener_ultimo_relevamiento, obtener_relevamiento_anterior

log = logging.getLogger(__name__)

REPORTS_DIR = Path(__file__).parent / "reports"

COLOR_HEADER_BG = "1F4E78"
COLOR_SUBE = "FFE5E5"   # rosa pálido para subas
COLOR_BAJA = "E5F7E5"   # verde pálido para bajas


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _agrupar_por_corte(precios: list[dict]) -> dict[str, dict[str, float]]:
    """
    Devuelve {corte: {carniceria: precio_kg_promedio}}.
    Si una carnicería tiene varios productos del mismo corte, promedia.
    """
    bucket: dict[str, dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))
    for p in precios:
        bucket[p["corte_normalizado"]][p["carniceria"]].append(p["precio_kg"])
    return {
        corte: {carn: round(mean(prs), 2) for carn, prs in carns.items()}
        for corte, carns in bucket.items()
    }


def _calcular_variaciones(actual: list[dict], anterior: list[dict]) -> dict[tuple[str, str], float]:
    """{(carniceria, corte): % variación vs anterior}"""
    ant_agrup = _agrupar_por_corte(anterior)
    act_agrup = _agrupar_por_corte(actual)
    variaciones = {}
    for corte, carns in act_agrup.items():
        for carn, precio_act in carns.items():
            precio_ant = ant_agrup.get(corte, {}).get(carn)
            if precio_ant and precio_ant > 0:
                pct = (precio_act - precio_ant) / precio_ant * 100
                variaciones[(carn, corte)] = round(pct, 1)
    return variaciones


# ─── Markdown ────────────────────────────────────────────────────────────────

def generar_markdown(actual: list[dict], anterior: list[dict],
                     fallidos: list, output_path: Path,
                     sospechosos: list | None = None):
    fecha_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    agrupado = _agrupar_por_corte(actual)
    variaciones = _calcular_variaciones(actual, anterior)
    carnicerias = sorted({p["carniceria"] for p in actual})

    out: list[str] = []
    out.append(f"# Relevamiento de precios de carne — {fecha_str}\n")
    out.append(f"- **Carnicerías:** {len(carnicerias)} ({', '.join(carnicerias)})")
    out.append(f"- **Cortes únicos relevados:** {len(agrupado)}")
    out.append(f"- **Productos totales:** {len(actual)}\n")

    if fallidos:
        out.append("## ⚠️ Scrapers con falla\n")
        for nombre, error in fallidos:
            out.append(f"- **{nombre}:** {error}")
        out.append("")

    if sospechosos:
        out.append("## ⚠️ Scrapers sospechosos (pocos cortes — ¿cambió el HTML?)\n")
        for nombre, n in sospechosos:
            out.append(f"- **{nombre}:** solo {n} cortes relevados")
        out.append("")

    out.append("## Tabla comparativa ($/kg)\n")
    header = ["Corte"] + carnicerias + ["Promedio", "Mín", "Máx"]
    out.append("| " + " | ".join(header) + " |")
    out.append("|" + "|".join(["---" if i == 0 else "---:" for i in range(len(header))]) + "|")

    for corte in sorted(agrupado.keys()):
        fila = [corte_pretty(corte)]
        precios_corte = []
        for carn in carnicerias:
            if carn in agrupado[corte]:
                p = agrupado[corte][carn]
                precios_corte.append(p)
                celda = f"${p:,.0f}".replace(",", ".")
                var = variaciones.get((carn, corte))
                if var is not None and abs(var) >= 5:
                    flecha = "🔺" if var > 0 else "🔻"
                    celda += f" {flecha}{abs(var):.0f}%"
                fila.append(celda)
            else:
                fila.append("—")
        if precios_corte:
            fila.append(f"${mean(precios_corte):,.0f}".replace(",", "."))
            fila.append(f"${min(precios_corte):,.0f}".replace(",", "."))
            fila.append(f"${max(precios_corte):,.0f}".replace(",", "."))
        else:
            fila.extend(["—", "—", "—"])
        out.append("| " + " | ".join(fila) + " |")

    # Variaciones destacadas
    out.append("\n## Variaciones significativas vs relevamiento anterior\n")
    sig = sorted(
        [(k, v) for k, v in variaciones.items() if abs(v) >= 5],
        key=lambda x: abs(x[1]),
        reverse=True,
    )[:15]
    if sig:
        out.append("| Carnicería | Corte | Variación |")
        out.append("|---|---|---:|")
        for (carn, corte), var in sig:
            flecha = "🔺" if var > 0 else "🔻"
            out.append(f"| {carn} | {corte_pretty(corte)} | {flecha} {var:+.1f}% |")
    else:
        out.append("_Sin variaciones >5% respecto al relevamiento anterior._")

    out.append("\n---\n_Generado automáticamente por relevador-carnes_")
    output_path.write_text("\n".join(out), encoding="utf-8")
    log.info(f"Markdown: {output_path}")


# ─── Excel ───────────────────────────────────────────────────────────────────

def generar_excel(actual: list[dict], anterior: list[dict], output_path: Path):
    agrupado = _agrupar_por_corte(actual)
    variaciones = _calcular_variaciones(actual, anterior)
    carnicerias = sorted({p["carniceria"] for p in actual})

    wb = openpyxl.Workbook()

    # Estilos comunes
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    bold = Font(bold=True)
    border = Border(*[Side(style="thin", color="DDDDDD")] * 4)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # === Hoja 1: Comparativa ===
    ws = wb.active
    ws.title = "Comparativa"

    titulo = f"Relevamiento de carnes — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(carnicerias) + 5)

    headers = ["Corte"] + carnicerias + ["Promedio", "Mín", "Máx", "Spread"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for r, corte in enumerate(sorted(agrupado.keys()), start=4):
        ws.cell(row=r, column=1, value=corte_pretty(corte)).font = bold
        ws.cell(row=r, column=1).border = border

        precios_corte = []
        for c, carn in enumerate(carnicerias, start=2):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = right
            if carn in agrupado[corte]:
                p = agrupado[corte][carn]
                precios_corte.append(p)
                cell.value = p
                cell.number_format = '"$"#,##0'
                var = variaciones.get((carn, corte))
                if var is not None:
                    if var >= 5:
                        cell.fill = PatternFill("solid", fgColor=COLOR_SUBE)
                    elif var <= -5:
                        cell.fill = PatternFill("solid", fgColor=COLOR_BAJA)

        base = len(carnicerias) + 2
        if precios_corte:
            ws.cell(row=r, column=base,     value=round(mean(precios_corte), 2)).number_format = '"$"#,##0'
            ws.cell(row=r, column=base + 1, value=round(min(precios_corte), 2)).number_format = '"$"#,##0'
            ws.cell(row=r, column=base + 2, value=round(max(precios_corte), 2)).number_format = '"$"#,##0'
            spread = (max(precios_corte) - min(precios_corte)) / min(precios_corte)
            ws.cell(row=r, column=base + 3, value=spread).number_format = '0.0%'
        for col_idx in range(base, base + 4):
            ws.cell(row=r, column=col_idx).border = border
            ws.cell(row=r, column=col_idx).alignment = right
            ws.cell(row=r, column=col_idx).font = bold

    # Anchos
    ws.column_dimensions["A"].width = 22
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.freeze_panes = "B4"

    # === Hoja 2: Detalle ===
    ws2 = wb.create_sheet("Detalle")
    det_headers = ["Fecha", "Carnicería", "Segmento", "Corte normalizado",
                   "Producto original", "Precio $/kg", "URL"]
    for col, h in enumerate(det_headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
    for r, p in enumerate(actual, start=2):
        ws2.cell(row=r, column=1, value=p["fecha"])
        ws2.cell(row=r, column=2, value=p["carniceria"])
        ws2.cell(row=r, column=3, value=p["segmento"])
        ws2.cell(row=r, column=4, value=corte_pretty(p["corte_normalizado"]))
        ws2.cell(row=r, column=5, value=p["corte_original"])
        ws2.cell(row=r, column=6, value=p["precio_kg"]).number_format = '"$"#,##0.00'
        ws2.cell(row=r, column=7, value=p["url_fuente"])

    for letra, w in zip("ABCDEFG", [12, 16, 14, 22, 42, 14, 60]):
        ws2.column_dimensions[letra].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    log.info(f"Excel: {output_path}")


# ─── PDF ─────────────────────────────────────────────────────────────────────

def generar_pdf(actual: list[dict], anterior: list[dict], output_path: Path,
                paths_graficos: dict[str, Path] | None = None):
    agrupado = _agrupar_por_corte(actual)
    variaciones = _calcular_variaciones(actual, anterior)
    carnicerias = sorted({p["carniceria"] for p in actual})

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=1.2 * cm, rightMargin=1.2 * cm,
        topMargin=1.2 * cm, bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=18, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9,
                               textColor=colors.grey, spaceAfter=10)
    h3 = ParagraphStyle("h3", parent=styles["Heading3"], fontSize=12, spaceBefore=12, spaceAfter=6)

    story = []
    fecha = datetime.now().strftime("%d/%m/%Y %H:%M")
    story.append(Paragraph("Relevamiento de precios de carne", title_style))
    story.append(Paragraph(
        f"Generado: {fecha} &nbsp;·&nbsp; {len(carnicerias)} carnicerías "
        f"&nbsp;·&nbsp; {len(agrupado)} cortes &nbsp;·&nbsp; {len(actual)} productos",
        sub_style
    ))

    headers = ["Corte"] + carnicerias + ["Prom.", "Mín.", "Máx."]
    rows = [headers]
    for corte in sorted(agrupado.keys()):
        fila = [corte_pretty(corte)]
        precios_corte = []
        for carn in carnicerias:
            if carn in agrupado[corte]:
                p = agrupado[corte][carn]
                precios_corte.append(p)
                txt = f"${p:,.0f}".replace(",", ".")
                var = variaciones.get((carn, corte))
                if var is not None and abs(var) >= 5:
                    txt += f"\n{var:+.0f}%"
                fila.append(txt)
            else:
                fila.append("—")
        if precios_corte:
            fila.append(f"${mean(precios_corte):,.0f}".replace(",", "."))
            fila.append(f"${min(precios_corte):,.0f}".replace(",", "."))
            fila.append(f"${max(precios_corte):,.0f}".replace(",", "."))
        else:
            fila.extend(["—", "—", "—"])
        rows.append(fila)

    col_widths = [3.5 * cm] + [2.6 * cm] * len(carnicerias) + [2.0 * cm] * 3
    table = Table(rows, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + COLOR_HEADER_BG)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
    ]))
    story.append(table)

    sig = sorted(
        [(k, v) for k, v in variaciones.items() if abs(v) >= 5],
        key=lambda x: abs(x[1]),
        reverse=True,
    )[:10]
    if sig:
        story.append(Paragraph("Variaciones destacadas vs relevamiento anterior", h3))
        var_rows = [["Carnicería", "Corte", "Variación"]]
        for (carn, corte), var in sig:
            var_rows.append([carn, corte_pretty(corte), f"{var:+.1f}%"])
        var_table = Table(var_rows, colWidths=[4 * cm, 5.5 * cm, 3 * cm])
        var_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + COLOR_HEADER_BG)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ]))
        story.append(var_table)

    # ─── Gráficos de tendencia ──────────────────────────────────────────
    if paths_graficos:
        story.append(PageBreak())
        story.append(Paragraph("Tendencias de precio (últimos 90 días)", h3))
        story.append(Spacer(1, 0.3 * cm))
        # Mostrar de a 2 gráficos por página, ordenados por nombre del corte
        for corte in sorted(paths_graficos.keys()):
            png_path = paths_graficos[corte]
            if png_path.exists():
                img = Image(str(png_path), width=24 * cm, height=12 * cm,
                            kind="proportional")
                story.append(img)
                story.append(Spacer(1, 0.3 * cm))

    doc.build(story)
    log.info(f"PDF: {output_path}")


# ─── Orquestador ─────────────────────────────────────────────────────────────

def generar_todos(fallidos: list | None = None,
                  paths_graficos: dict[str, Path] | None = None,
                  sospechosos: list | None = None) -> dict[str, Path]:
    fallidos = fallidos or []
    sospechosos = sospechosos or []
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    actual = obtener_ultimo_relevamiento()
    anterior = obtener_relevamiento_anterior()

    if not actual:
        raise RuntimeError(
            "No hay datos en la base. Corré primero `python run.py` "
            "para hacer el primer relevamiento."
        )

    fecha_str = datetime.now().strftime("%Y-%m-%d_%H%M")
    paths = {
        "md":   REPORTS_DIR / f"reporte_{fecha_str}.md",
        "xlsx": REPORTS_DIR / f"reporte_{fecha_str}.xlsx",
        "pdf":  REPORTS_DIR / f"reporte_{fecha_str}.pdf",
    }

    generar_markdown(actual, anterior, fallidos, paths["md"], sospechosos=sospechosos)
    generar_excel(actual, anterior, paths["xlsx"])
    generar_pdf(actual, anterior, paths["pdf"], paths_graficos=paths_graficos)

    return paths


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    paths = generar_todos()
    print("Reportes generados:")
    for k, p in paths.items():
        print(f"  {k:5s} -> {p}")
